#!/usr/bin/env python3
"""
Excel Splitter for SharePoint
將 Excel 檔案依據 Approver 欄位拆分成多個子檔案
"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path


def find_approver_column(worksheet):
    """找出 Approver 欄位的位置"""
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == 'Approver':
            return col_idx
    raise ValueError("找不到 'Approver' 欄位！請確認欄位名稱")


def split_excel_by_approver(file_path):
    """主要處理函數：讀取 Excel 並按 Approver 分檔"""
    
    # 檢查檔案存在
    if not os.path.exists(file_path):
        print(f"錯誤：找不到檔案 {file_path}")
        sys.exit(1)
    
    # 讀取 Excel 取得所有 Approver
    print(f"讀取檔案: {file_path}")
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
    except Exception as e:
        print(f"讀取 Excel 失敗: {e}")
        sys.exit(1)
    
    # 檢查 Approver 欄位存在
    if 'Approver' not in df.columns:
        print("錯誤：Excel 中找不到 'Approver' 欄位")
        sys.exit(1)
    
    # 取得所有唯一的 Approver
    approvers = df['Approver'].dropna().unique().tolist()
    print(f"找到 {len(approvers)} 位 Approver: {', '.join(str(a) for a in approvers)}")
    
    # 取得基礎路徑和檔名
    base_dir = os.path.dirname(file_path)
    base_name = os.path.basename(file_path)
    
    # 為每個 Approver 建立資料夾並複製檔案
    for approver in approvers:
        approver_name = str(approver).strip()
        
        # 建立子資料夾
        subdir = os.path.join(base_dir, approver_name)
        os.makedirs(subdir, exist_ok=True)
        
        # 目標檔案路徑
        dst_path = os.path.join(subdir, base_name)
        
        # 載入原始 Excel 檔案
        wb = load_workbook(file_path)
        ws = wb.active
        
        try:
            # 找出 Approver 欄位位置
            approver_col = find_approver_column(ws)
            
            # 取得資料範圍
            max_row = ws.max_row
            max_col = ws.max_column
            
            # 設定 AutoFilter 範圍
            filter_range = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.auto_filter.ref = filter_range
            
            # 套用篩選條件 - 只顯示該 Approver 的資料
            # 注意：openpyxl 的 column index 是從 0 開始
            ws.auto_filter.add_filter_column(approver_col - 1, [approver_name])
            
            # 儲存檔案
            wb.save(dst_path)
            print(f"✓ 已建立 {approver_name} 的檔案: {dst_path}")
            
        except Exception as e:
            print(f"✗ 處理 {approver_name} 時發生錯誤: {e}")
        finally:
            wb.close()
    
    print("\n處理完成！")
    print(f"所有檔案都已建立在原始檔案的同一層目錄下")


def main():
    """主程式進入點"""
    if len(sys.argv) != 2:
        print("使用方式: python splitter.py <Excel檔案路徑>")
        print("範例: python splitter.py /path/to/master.xlsx")
        sys.exit(1)
    
    file_path = sys.argv[1]
    split_excel_by_approver(file_path)


if __name__ == "__main__":
    main()