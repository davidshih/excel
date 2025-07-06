#!/usr/bin/env python3

import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path


def find_approver_column(worksheet):
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == 'Reviewer':
            return col_idx
    raise ValueError("Cannot find 'Reviewer' column! Please check column name")


def split_excel_by_approver(file_path):
    if not os.path.exists(file_path):
        print(f"Error: File not found {file_path}")
        sys.exit(1)
    
    print(f"Reading file: {file_path}")
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
    except Exception as e:
        print(f"Failed to read Excel: {e}")
        sys.exit(1)
    
    if 'Reviewer' not in df.columns:
        print("Error: Cannot find 'Reviewer' column in Excel")
        sys.exit(1)
    
    approvers = df['Reviewer'].dropna().unique().tolist()
    print(f"Found {len(approvers)} reviewers: {', '.join(str(a) for a in approvers)}")
    
    base_dir = os.path.dirname(file_path)
    base_name = os.path.basename(file_path)
    
    for approver in approvers:
        approver_name = str(approver).strip()
        
        subdir = os.path.join(base_dir, approver_name)
        os.makedirs(subdir, exist_ok=True)
        
        dst_path = os.path.join(subdir, base_name)
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        try:
            approver_col = find_approver_column(ws)
            
            max_row = ws.max_row
            max_col = ws.max_column
            
            filter_range = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.auto_filter.ref = filter_range
            
            ws.auto_filter.add_filter_column(approver_col - 1, [approver_name])
            
            wb.save(dst_path)
            print(f"✓ Created file for {approver_name}: {dst_path}")
            
        except Exception as e:
            print(f"✗ Error processing {approver_name}: {e}")
        finally:
            wb.close()
    
    print("\nProcessing complete!")
    print(f"All files have been created in the same directory as the original file")


def main():
    if len(sys.argv) != 2:
        print("Usage: python splitter_en.py <Excel file path>")
        print("Example: python splitter_en.py /path/to/master.xlsx")
        sys.exit(1)
    
    file_path = sys.argv[1]
    split_excel_by_approver(file_path)


if __name__ == "__main__":
    main()