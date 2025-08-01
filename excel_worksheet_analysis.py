#!/usr/bin/env python3
"""
Excel 工作表分析工具
用於診斷資料驗證跨工作表引用問題

功能：
1. 分析所有工作表
2. 檢測資料驗證規則
3. 找出跨工作表依賴關係
4. 提供修復建議
"""

import os
import sys
from typing import Dict, List, Set, Tuple, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("❌ 請安裝 openpyxl: pip install openpyxl")
    sys.exit(1)

if sys.platform == 'win32':
    try:
        import win32com.client
        WIN32COM_AVAILABLE = True
    except ImportError:
        WIN32COM_AVAILABLE = False
else:
    WIN32COM_AVAILABLE = False


class ExcelWorksheetAnalyzer:
    """Excel 工作表分析器"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.analysis_result = {}
    
    def analyze_with_openpyxl(self) -> Dict:
        """使用 openpyxl 分析工作表"""
        print("🔍 使用 openpyxl 分析工作表...")
        
        try:
            self.workbook = load_workbook(self.file_path, data_only=False)
            result = {
                'worksheets': {},
                'data_validations': {},
                'cross_sheet_references': [],
                'summary': {}
            }
            
            # 分析每個工作表
            for ws_name in self.workbook.sheetnames:
                ws = self.workbook[ws_name]
                
                # 基本資訊
                ws_info = {
                    'name': ws_name,
                    'max_row': ws.max_row,
                    'max_column': ws.max_column,
                    'data_validations': [],
                    'named_ranges': [],
                    'has_data': ws.max_row > 1
                }
                
                # 檢查資料驗證
                if hasattr(ws, 'data_validations') and ws.data_validations:
                    for dv in ws.data_validations.dataValidation:
                        dv_info = self.analyze_data_validation(dv, ws_name)
                        ws_info['data_validations'].append(dv_info)
                        
                        # 檢查跨工作表引用
                        if dv_info['has_cross_sheet_reference']:
                            result['cross_sheet_references'].append({
                                'source_sheet': ws_name,
                                'validation_info': dv_info
                            })
                
                result['worksheets'][ws_name] = ws_info
            
            # 生成摘要
            result['summary'] = {
                'total_worksheets': len(result['worksheets']),
                'worksheets_with_data': sum(1 for ws in result['worksheets'].values() if ws['has_data']),
                'total_data_validations': sum(len(ws['data_validations']) for ws in result['worksheets'].values()),
                'cross_sheet_references': len(result['cross_sheet_references'])
            }
            
            return result
            
        except Exception as e:
            print(f"❌ openpyxl 分析失敗: {e}")
            return {}
        
        finally:
            if self.workbook:
                self.workbook.close()
    
    def analyze_data_validation(self, dv: DataValidation, sheet_name: str) -> Dict:
        """分析資料驗證規則"""
        dv_info = {
            'type': str(dv.type),
            'formula1': str(dv.formula1) if dv.formula1 else None,
            'formula2': str(dv.formula2) if dv.formula2 else None,
            'ranges': [str(range_obj) for range_obj in dv.ranges],
            'has_cross_sheet_reference': False,
            'referenced_sheets': []
        }
        
        # 檢查公式中的跨工作表引用
        for formula in [dv.formula1, dv.formula2]:
            if formula:
                formula = str(formula)
                if '!' in formula:
                    dv_info['has_cross_sheet_reference'] = True
                    # 提取工作表名稱
                    parts = formula.split('!')
                    for part in parts[:-1]:
                        # 清理工作表名稱
                        sheet_ref = part.split("'")[-1] if "'" in part else part
                        if sheet_ref and sheet_ref != sheet_name:
                            dv_info['referenced_sheets'].append(sheet_ref)
        
        return dv_info
    
    def analyze_with_com(self) -> Dict:
        """使用 COM 分析工作表（更詳細）"""
        if not WIN32COM_AVAILABLE:
            print("⚠️ COM 不可用，跳過 COM 分析")
            return {}
        
        print("🔍 使用 COM 分析工作表...")
        excel = None
        wb = None
        
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            
            wb = excel.Workbooks.Open(os.path.abspath(self.file_path))
            
            result = {
                'worksheets': {},
                'named_ranges': {},
                'summary': {}
            }
            
            # 分析工作表
            for i in range(1, wb.Worksheets.Count + 1):
                ws = wb.Worksheets(i)
                ws_name = ws.Name
                
                ws_info = {
                    'name': ws_name,
                    'used_range': str(ws.UsedRange.Address),
                    'visible': ws.Visible,
                    'data_validations': []
                }
                
                # 檢查資料驗證（COM 方式）
                try:
                    used_range = ws.UsedRange
                    for row in range(1, used_range.Rows.Count + 1):
                        for col in range(1, used_range.Columns.Count + 1):
                            cell = ws.Cells(row, col)
                            if hasattr(cell, 'Validation') and cell.Validation.Type != 0:  # 0 = xlValidateInputOnly
                                validation_info = {
                                    'cell': f"{ws_name}!{cell.Address}",
                                    'type': cell.Validation.Type,
                                    'formula1': cell.Validation.Formula1 if hasattr(cell.Validation, 'Formula1') else None
                                }
                                ws_info['data_validations'].append(validation_info)
                except:
                    pass  # 有些儲存格可能無法存取
                
                result['worksheets'][ws_name] = ws_info
            
            # 分析命名範圍
            try:
                for i in range(1, wb.Names.Count + 1):
                    name = wb.Names(i)
                    result['named_ranges'][name.Name] = {
                        'refers_to': name.RefersTo,
                        'visible': name.Visible
                    }
            except:
                pass
            
            return result
            
        except Exception as e:
            print(f"❌ COM 分析失敗: {e}")
            return {}
        
        finally:
            if wb:
                try:
                    wb.Close(SaveChanges=False)
                except:
                    pass
            if excel:
                try:
                    excel.Quit()
                except:
                    pass
    
    def generate_report(self, analysis: Dict) -> str:
        """生成分析報告"""
        if not analysis:
            return "❌ 無法生成報告：分析資料為空"
        
        report = []
        report.append("📊 Excel 工作表分析報告")
        report.append("=" * 50)
        
        # 摘要
        if 'summary' in analysis:
            summary = analysis['summary']
            report.append(f"📋 摘要資訊：")
            report.append(f"  • 總工作表數：{summary.get('total_worksheets', 0)}")
            report.append(f"  • 含資料工作表：{summary.get('worksheets_with_data', 0)}")
            report.append(f"  • 資料驗證規則：{summary.get('total_data_validations', 0)}")
            report.append(f"  • 跨工作表引用：{summary.get('cross_sheet_references', 0)}")
            report.append("")
        
        # 工作表詳情
        if 'worksheets' in analysis:
            report.append("📄 工作表詳情：")
            for ws_name, ws_info in analysis['worksheets'].items():
                report.append(f"  🗂️ {ws_name}")
                report.append(f"    - 資料範圍：{ws_info.get('max_row', 0)} 行 x {ws_info.get('max_column', 0)} 欄")
                report.append(f"    - 含資料：{'是' if ws_info.get('has_data', False) else '否'}")
                
                if ws_info.get('data_validations'):
                    report.append(f"    - 資料驗證：{len(ws_info['data_validations'])} 個規則")
                    for dv in ws_info['data_validations']:
                        if dv.get('has_cross_sheet_reference'):
                            report.append(f"      ⚠️ 跨工作表引用：{dv.get('referenced_sheets', [])}")
                report.append("")
        
        # 跨工作表引用警告
        if 'cross_sheet_references' in analysis and analysis['cross_sheet_references']:
            report.append("⚠️ 跨工作表引用警告：")
            for ref in analysis['cross_sheet_references']:
                source = ref['source_sheet']
                validation = ref['validation_info']
                referenced = validation.get('referenced_sheets', [])
                report.append(f"  • {source} → {', '.join(referenced)}")
                report.append(f"    公式：{validation.get('formula1', 'N/A')}")
            report.append("")
        
        # 建議
        report.append("💡 建議：")
        if analysis.get('cross_sheet_references'):
            report.append("  ⚠️ 發現跨工作表引用，處理時需要：")
            report.append("    1. 保留所有被引用的工作表")
            report.append("    2. 只對主要資料工作表套用篩選")
            report.append("    3. 使用隱藏列而非刪除列")
        else:
            report.append("  ✅ 沒有發現跨工作表引用，可以安全處理")
        
        return "\n".join(report)
    
    def full_analysis(self) -> str:
        """完整分析並生成報告"""
        print(f"📁 分析檔案：{os.path.basename(self.file_path)}")
        
        # 使用 openpyxl 分析
        openpyxl_result = self.analyze_with_openpyxl()
        
        # 如果可用，也使用 COM 分析
        com_result = self.analyze_with_com() if WIN32COM_AVAILABLE else {}
        
        # 合併結果（以 openpyxl 為主）
        final_result = openpyxl_result
        
        # 生成報告
        report = self.generate_report(final_result)
        
        return report


def analyze_excel_file(file_path: str) -> str:
    """分析 Excel 檔案的工作表結構"""
    if not os.path.exists(file_path):
        return f"❌ 檔案不存在：{file_path}"
    
    analyzer = ExcelWorksheetAnalyzer(file_path)
    return analyzer.full_analysis()


def main():
    """主程式"""
    if len(sys.argv) != 2:
        print("使用方式：python excel_worksheet_analysis.py <excel_file_path>")
        print("範例：python excel_worksheet_analysis.py test.xlsx")
        return
    
    file_path = sys.argv[1]
    report = analyze_excel_file(file_path)
    print(report)
    
    # 儲存報告
    report_path = f"{os.path.splitext(file_path)[0]}_analysis_report.txt"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n📄 報告已儲存至：{report_path}")


if __name__ == "__main__":
    main()