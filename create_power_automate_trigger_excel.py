#!/usr/bin/env python3
"""
建立 Power Automate 觸發用的 Excel 檔案
這個檔案包含一個表格，Power Automate 會監控這個表格的變化
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_power_automate_trigger_excel(output_path: str = "power_automate_permissions_trigger.xlsx"):
    """
    建立 Power Automate 使用的 Excel 觸發檔案
    
    Args:
        output_path: 輸出檔案路徑
    """
    print("🚀 建立 Power Automate 觸發用 Excel 檔案...")
    
    # 建立範例資料
    sample_data = [
        {
            '審查者名稱': '張三',
            '資料夾名稱': '張三',
            '資料夾完整路徑': '/sites/YourSite/Documents/張三',
            'Email': 'zhangsan@company.com',
            '權限等級': 'Contribute',
            '處理狀態': '待處理',
            '處理時間': '',
            '處理結果': '',
            '站台名稱': 'https://company.sharepoint.com/sites/YourSite',
            '文件庫': 'Documents'
        },
        {
            '審查者名稱': '李四',
            '資料夾名稱': '李四', 
            '資料夾完整路徑': '/sites/YourSite/Documents/李四',
            'Email': 'lisi@company.com',
            '權限等級': 'Contribute',
            '處理狀態': '待處理',
            '處理時間': '',
            '處理結果': '',
            '站台名稱': 'https://company.sharepoint.com/sites/YourSite',
            '文件庫': 'Documents'
        }
    ]
    
    # 建立 DataFrame
    df = pd.DataFrame(sample_data)
    
    # 建立 Excel Writer
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # 寫入主要資料表
        df.to_excel(writer, sheet_name='權限設定', index=False)
        
        # 取得工作簿和工作表
        workbook = writer.book
        worksheet = workbook['權限設定']
        
        # 設定欄位寬度
        column_widths = {
            'A': 15,  # 審查者名稱
            'B': 15,  # 資料夾名稱
            'C': 40,  # 資料夾完整路徑
            'D': 30,  # Email
            'E': 15,  # 權限等級
            'F': 12,  # 處理狀態
            'G': 20,  # 處理時間
            'H': 30,  # 處理結果
            'I': 50,  # 站台名稱
            'J': 15   # 文件庫
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 設定標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 建立表格（Excel Table）
        tab = Table(displayName="權限設定表", ref=f"A1:J{len(df)+1}")
        
        # 設定表格樣式
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        
        # 新增說明工作表
        instruction_sheet = workbook.create_sheet("使用說明")
        
        instructions = [
            ["Power Automate SharePoint 權限設定工具使用說明", ""],
            ["", ""],
            ["步驟 1：設定 Power Automate", ""],
            ["1. 登入 Power Automate (https://make.powerautomate.com/)", ""],
            ["2. 匯入提供的 power_automate_sharepoint_permissions.json 檔案", ""],
            ["3. 更新流程中的 EXCEL_FILE_ID 為此檔案的 ID", ""],
            ["4. 設定連線（Excel Online Business、SharePoint、Office 365）", ""],
            ["", ""],
            ["步驟 2：使用此 Excel 檔案", ""],
            ["1. 將此檔案上傳到 OneDrive 或 SharePoint", ""],
            ["2. 在「權限設定」工作表中填入資料：", ""],
            ["   - 審查者名稱：使用者的顯示名稱", ""],
            ["   - 資料夾名稱：要授權的資料夾名稱", ""],
            ["   - Email：使用者的電子郵件地址", ""],
            ["   - 權限等級：Contribute（參與）或 Read（讀取）", ""],
            ["   - 站台名稱：完整的 SharePoint 網站 URL", ""],
            ["   - 文件庫：文件庫名稱（通常是 Documents）", ""],
            ["", ""],
            ["步驟 3：觸發處理", ""],
            ["1. 將「處理狀態」設為「待處理」", ""],
            ["2. Power Automate 會自動偵測並處理（每 5 分鐘檢查一次）", ""],
            ["3. 處理完成後，狀態會更新為「已處理」或「處理失敗」", ""],
            ["", ""],
            ["權限等級說明：", ""],
            ["- Read：僅可檢視和下載檔案", ""],
            ["- Contribute：可檢視、下載、上傳、編輯檔案", ""],
            ["- Edit：可檢視、下載、上傳、編輯、刪除檔案", ""],
            ["- Full Control：完全控制（請謹慎使用）", ""],
            ["", ""],
            ["注意事項：", ""],
            ["1. 確保 Email 地址正確且使用者存在於組織中", ""],
            ["2. 資料夾必須已存在於指定的文件庫中", ""],
            ["3. 執行者需要有足夠的權限來設定資料夾權限", ""],
            ["4. 建議先在測試環境中測試流程", ""]
        ]
        
        # 寫入說明
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = instruction_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # 設定標題樣式
                if row_idx == 1 and col_idx == 1:
                    cell.font = Font(bold=True, size=14, color="366092")
                elif value.startswith("步驟") or value.endswith("說明：") or value == "注意事項：":
                    cell.font = Font(bold=True, size=12)
        
        # 設定說明工作表的欄寬
        instruction_sheet.column_dimensions['A'].width = 80
        instruction_sheet.column_dimensions['B'].width = 40
        
        # 建立資料驗證工作表
        validation_sheet = workbook.create_sheet("資料驗證")
        
        # 權限等級選項
        permission_levels = ["Read", "Contribute", "Edit", "Full Control"]
        validation_sheet.append(["權限等級選項"])
        for level in permission_levels:
            validation_sheet.append([level])
        
        # 處理狀態選項
        validation_sheet.append([])
        validation_sheet.append(["處理狀態選項"])
        status_options = ["待處理", "處理中", "已處理", "處理失敗"]
        for status in status_options:
            validation_sheet.append([status])
        
        # 隱藏資料驗證工作表
        validation_sheet.sheet_state = 'hidden'
    
    print(f"✅ Excel 觸發檔案已建立：{output_path}")
    print("\n📋 下一步：")
    print("1. 將此檔案上傳到 OneDrive 或 SharePoint")
    print("2. 在 Power Automate 中更新檔案 ID")
    print("3. 開始使用！")
    
    return output_path


def create_batch_import_template(
    processed_folders: list,
    output_path: str = "batch_permissions_import.xlsx"
):
    """
    建立批次匯入模板，用於大量設定權限
    
    Args:
        processed_folders: 已處理的資料夾清單
        output_path: 輸出檔案路徑
    """
    print(f"\n🚀 建立批次權限匯入檔案...")
    
    # 準備資料
    data = []
    for folder_info in processed_folders:
        data.append({
            '審查者名稱': folder_info['reviewer'],
            '資料夾名稱': folder_info['folder_name'],
            '資料夾完整路徑': folder_info['full_path'],
            'Email': f"{folder_info['reviewer'].lower().replace(' ', '.')}@company.com",  # 預設格式
            '權限等級': 'Contribute',
            '處理狀態': '待處理',
            '處理時間': '',
            '處理結果': '',
            '站台名稱': 'https://company.sharepoint.com/sites/YourSite',  # 需要修改
            '文件庫': 'Documents'
        })
    
    # 建立 DataFrame 並儲存
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='批次權限設定', index=False)
        
        # 設定格式
        workbook = writer.book
        worksheet = workbook['批次權限設定']
        
        # 建立表格
        if len(df) > 0:
            tab = Table(displayName="批次權限表", ref=f"A1:J{len(df)+1}")
            style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            worksheet.add_table(tab)
    
    print(f"✅ 批次匯入檔案已建立：{output_path}")
    print(f"📊 包含 {len(data)} 個資料夾的權限設定")
    
    return output_path


if __name__ == "__main__":
    # 建立觸發檔案
    trigger_file = create_power_automate_trigger_excel()
    
    # 建立批次匯入範例
    sample_folders = [
        {'reviewer': '張三', 'folder_name': '張三', 'full_path': '/Documents/張三'},
        {'reviewer': '李四', 'folder_name': '李四', 'full_path': '/Documents/李四'},
        {'reviewer': '王五', 'folder_name': '王五', 'full_path': '/Documents/王五'}
    ]
    
    batch_file = create_batch_import_template(sample_folders)
    
    print("\n✨ 所有檔案已建立完成！")