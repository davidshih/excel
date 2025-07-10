#!/usr/bin/env python3

import sys
import os
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import glob
import argparse


def find_column(worksheet, column_name):
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == column_name:
            return col_idx
    raise ValueError(f"Cannot find '{column_name}' column! Please check column name")


def copy_documents(source_dir, dest_dir, app_name):
    word_pattern = os.path.join(source_dir, f"{app_name}*.docx")
    word_files = glob.glob(word_pattern)
    
    pdf_pattern = os.path.join(source_dir, f"{app_name}*permission*.pdf")
    pdf_files = glob.glob(pdf_pattern)
    
    # For testing, also look for .txt files
    if not word_files:
        word_pattern = os.path.join(source_dir, f"{app_name}*.txt")
        word_files = [f for f in glob.glob(word_pattern) if "permission" not in f]
    
    if not pdf_files:
        pdf_pattern = os.path.join(source_dir, f"{app_name}*permission*.txt")
        pdf_files = glob.glob(pdf_pattern)
    
    copied_files = []
    
    for word_file in word_files:
        dest_path = os.path.join(dest_dir, os.path.basename(word_file))
        shutil.copy2(word_file, dest_path)
        copied_files.append(os.path.basename(word_file))
    
    for pdf_file in pdf_files:
        dest_path = os.path.join(dest_dir, os.path.basename(pdf_file))
        shutil.copy2(pdf_file, dest_path)
        copied_files.append(os.path.basename(pdf_file))
    
    return copied_files


def create_sharepoint_sharing_script(base_dir, reviewer_emails):
    script_path = os.path.join(base_dir, "share_folders.ps1")
    
    with open(script_path, 'w', encoding='utf-8') as f:
        f.write("# PowerShell script to share folders on SharePoint\n")
        f.write("# Run this script after uploading folders to SharePoint\n\n")
        f.write("$siteUrl = Read-Host 'Enter SharePoint site URL'\n")
        f.write("$baseFolder = Read-Host 'Enter base folder path on SharePoint'\n\n")
        f.write("Connect-PnPOnline -Url $siteUrl -UseWebLogin\n\n")
        
        for reviewer_name, email in reviewer_emails.items():
            f.write(f"# Share folder for {reviewer_name}\n")
            f.write(f"$folderPath = Join-Path $baseFolder '{reviewer_name}'\n")
            
            if email and email != 'N/A':
                f.write(f"$userEmail = '{email}'\n")
                f.write(f"Write-Host 'Sharing with {reviewer_name} ({email})...'\n")
            else:
                f.write(f"$userEmail = Read-Host 'Enter email for {reviewer_name}'\n")
            
            f.write(f"try {{\n")
            f.write(f"    Set-PnPFolderPermission -List 'Documents' -Identity $folderPath -User $userEmail -AddRole 'Edit'\n")
            f.write(f"    Write-Host '✓ Shared folder for {reviewer_name} with Edit permissions' -ForegroundColor Green\n")
            f.write(f"}} catch {{\n")
            f.write(f"    Write-Host '✗ Failed to share with {reviewer_name}: $_' -ForegroundColor Red\n")
            f.write(f"}}\n\n")
    
    return script_path


def split_excel_enhanced(file_path, app_name):
    if not os.path.exists(file_path):
        print(f"Error: File not found {file_path}")
        sys.exit(1)
    
    print(f"Reading file: {file_path}")
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
    except Exception as e:
        print(f"Failed to read Excel: {e}")
        sys.exit(1)
    
    if 'Reviewer' not in df.columns:
        print("Error: Cannot find 'Reviewer' column in Excel")
        sys.exit(1)
    
    reviewers = df['Reviewer'].dropna().unique().tolist()
    print(f"Found {len(reviewers)} reviewers: {', '.join(str(r) for r in reviewers)}")
    
    # Check for Email Address column and create mapping
    reviewer_emails = {}
    if 'Email Address' in df.columns:
        print("✓ Found 'Email Address' column - will use for automatic sharing")
        for reviewer in reviewers:
            reviewer_data = df[df['Reviewer'] == reviewer]
            if not reviewer_data.empty:
                email = reviewer_data['Email Address'].iloc[0]
                if pd.notna(email):
                    reviewer_emails[str(reviewer).strip()] = str(email).strip()
                else:
                    reviewer_emails[str(reviewer).strip()] = 'N/A'
    else:
        print("ℹ No 'Email Address' column found - will prompt for emails during sharing")
        reviewer_emails = {str(r).strip(): 'N/A' for r in reviewers}
    
    base_dir = os.path.dirname(file_path)
    app_folder = os.path.join(base_dir, app_name)
    os.makedirs(app_folder, exist_ok=True)
    print(f"Created application folder: {app_folder}")
    
    base_name = os.path.basename(file_path)
    
    for reviewer in reviewers:
        reviewer_name = str(reviewer).strip()
        
        reviewer_folder = os.path.join(app_folder, reviewer_name)
        os.makedirs(reviewer_folder, exist_ok=True)
        
        dst_path = os.path.join(reviewer_folder, base_name)
        
        wb = load_workbook(file_path)
        ws = wb.active
        
        try:
            reviewer_col = find_column(ws, 'Reviewer')
            
            max_row = ws.max_row
            max_col = ws.max_column
            
            filter_range = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.auto_filter.ref = filter_range
            
            ws.auto_filter.add_filter_column(reviewer_col - 1, [reviewer_name])
            
            # 隱藏不屬於此 reviewer 的資料行
            for row in range(2, max_row + 1):  # 從第2行開始（跳過標題）
                cell_value = ws.cell(row=row, column=reviewer_col).value
                if cell_value != reviewer_name:
                    ws.row_dimensions[row].hidden = True
            
            wb.save(dst_path)
            print(f"✓ Created filtered Excel for {reviewer_name}")
            
            copied_docs = copy_documents(base_dir, reviewer_folder, app_name)
            if copied_docs:
                print(f"  ✓ Copied documents: {', '.join(copied_docs)}")
            
        except Exception as e:
            print(f"✗ Error processing {reviewer_name}: {e}")
        finally:
            wb.close()
    
    script_path = create_sharepoint_sharing_script(app_folder, reviewer_emails)
    print(f"\n✓ Created SharePoint sharing script: {script_path}")
    
    # Display email mapping summary
    if any(email != 'N/A' for email in reviewer_emails.values()):
        print("\n📧 Email addresses found:")
        for reviewer, email in reviewer_emails.items():
            if email != 'N/A':
                print(f"  • {reviewer}: {email}")
            else:
                print(f"  • {reviewer}: [No email - will be prompted]")
    
    print("\nProcessing complete!")
    print(f"All files have been created in: {app_folder}")
    print("\nNext steps:")
    print("1. Upload the entire folder structure to SharePoint")
    print("2. Run the PowerShell script 'share_folders.ps1' to set permissions")
    print("3. The script will prompt for reviewer email addresses")


def main():
    parser = argparse.ArgumentParser(description='Split Excel by reviewer with enhanced features')
    parser.add_argument('excel_file', help='Path to the Excel file')
    parser.add_argument('app_name', help='Application name for the main folder')
    
    args = parser.parse_args()
    
    split_excel_enhanced(args.excel_file, args.app_name)


if __name__ == "__main__":
    main()