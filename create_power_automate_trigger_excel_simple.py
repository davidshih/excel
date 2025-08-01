#!/usr/bin/env python3
"""
建立 Power Automate 觸發用的 Excel 檔案 (簡化版)
使用 openpyxl 直接建立，不依賴 pandas
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


def create_power_automate_trigger_excel(output_path: str = "power_automate_permissions_trigger.xlsx"):
    """
    建立 Power Automate 使用的 Excel 觸發檔案
    
    Args:
        output_path: 輸出檔案路徑
    """
    print("🚀 建立 Power Automate 觸發用 Excel 檔案...")
    
    # 建立新的工作簿
    wb = Workbook()
    
    # 取得預設工作表並重新命名
    ws = wb.active
    ws.title = "權限設定"
    
    # 定義欄位標題
    headers = [
        '審查者名稱', '資料夾名稱', '資料夾完整路徑', 'Email', 
        '權限等級', '處理狀態', '處理時間', '處理結果', 
        '站台名稱', '文件庫'
    ]
    
    # 寫入標題
    ws.append(headers)
    
    # 建立範例資料
    sample_data = [
        ['張三', '張三', '/sites/YourSite/Documents/張三', 'zhangsan@company.com', 
         'Contribute', '待處理', '', '', 
         'https://company.sharepoint.com/sites/YourSite', 'Documents'],
        ['李四', '李四', '/sites/YourSite/Documents/李四', 'lisi@company.com', 
         'Contribute', '待處理', '', '', 
         'https://company.sharepoint.com/sites/YourSite', 'Documents']
    ]
    
    # 寫入範例資料
    for row in sample_data:
        ws.append(row)
    
    # 設定欄位寬度
    column_widths = {
        'A': 15,  # 審查者名稱
        'B': 15,  # 資料夾名稱
        'C': 40,  # 資料夾完整路徑
        'D': 30,  # Email
        'E': 15,  # 權限等級
        'F': 12,  # 處理狀態
        'G': 20,  # 處理時間
        'H': 30,  # 處理結果
        'I': 50,  # 站台名稱
        'J': 15   # 文件庫
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 設定標題樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 建立表格（Excel Table）
    tab = Table(displayName="權限設定表", ref=f"A1:J{ws.max_row}")
    
    # 設定表格樣式
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    
    # 新增資料驗證工作表
    validation_ws = wb.create_sheet("資料驗證")
    
    # 權限等級選項
    validation_ws.append(["權限等級選項"])
    permission_levels = ["Read", "Contribute", "Edit", "Full Control"]
    for i, level in enumerate(permission_levels, 2):
        validation_ws.cell(row=i, column=1, value=level)
    
    # 處理狀態選項
    validation_ws.cell(row=7, column=1, value="處理狀態選項")
    status_options = ["待處理", "處理中", "已處理", "處理失敗"]
    for i, status in enumerate(status_options, 8):
        validation_ws.cell(row=i, column=1, value=status)
    
    # 建立資料驗證 - 權限等級
    dv_permission = DataValidation(
        type="list",
        formula1="=資料驗證!$A$2:$A$5",
        allow_blank=False,
        showDropDown=True,
        showErrorMessage=True,
        errorTitle="無效的權限等級",
        error="請從下拉選單中選擇有效的權限等級"
    )
    ws.add_data_validation(dv_permission)
    dv_permission.add(f"E2:E{ws.max_row}")
    
    # 建立資料驗證 - 處理狀態
    dv_status = DataValidation(
        type="list",
        formula1="=資料驗證!$A$8:$A$11",
        allow_blank=False,
        showDropDown=True,
        showErrorMessage=True,
        errorTitle="無效的處理狀態",
        error="請從下拉選單中選擇有效的處理狀態"
    )
    ws.add_data_validation(dv_status)
    dv_status.add(f"F2:F{ws.max_row}")
    
    # 隱藏資料驗證工作表
    validation_ws.sheet_state = 'hidden'
    
    # 新增說明工作表
    instruction_ws = wb.create_sheet("使用說明")
    
    instructions = [
        ["Power Automate SharePoint 權限設定工具使用說明"],
        [""],
        ["步驟 1：設定 Power Automate"],
        ["1. 登入 Power Automate (https://make.powerautomate.com/)"],
        ["2. 匯入提供的 power_automate_sharepoint_permissions.json 檔案"],
        ["3. 更新流程中的 EXCEL_FILE_ID 為此檔案的 ID"],
        ["4. 設定連線（Excel Online Business、SharePoint、Office 365）"],
        [""],
        ["步驟 2：使用此 Excel 檔案"],
        ["1. 將此檔案上傳到 OneDrive 或 SharePoint"],
        ["2. 在「權限設定」工作表中填入資料："],
        ["   - 審查者名稱：使用者的顯示名稱"],
        ["   - 資料夾名稱：要授權的資料夾名稱"],
        ["   - Email：使用者的電子郵件地址"],
        ["   - 權限等級：Contribute（參與）或 Read（讀取）"],
        ["   - 站台名稱：完整的 SharePoint 網站 URL"],
        ["   - 文件庫：文件庫名稱（通常是 Documents）"],
        [""],
        ["步驟 3：觸發處理"],
        ["1. 將「處理狀態」設為「待處理」"],
        ["2. Power Automate 會自動偵測並處理（每 5 分鐘檢查一次）"],
        ["3. 處理完成後，狀態會更新為「已處理」或「處理失敗」"],
        [""],
        ["權限等級說明："],
        ["- Read：僅可檢視和下載檔案"],
        ["- Contribute：可檢視、下載、上傳、編輯檔案"],
        ["- Edit：可檢視、下載、上傳、編輯、刪除檔案"],
        ["- Full Control：完全控制（請謹慎使用）"],
        [""],
        ["注意事項："],
        ["1. 確保 Email 地址正確且使用者存在於組織中"],
        ["2. 資料夾必須已存在於指定的文件庫中"],
        ["3. 執行者需要有足夠的權限來設定資料夾權限"],
        ["4. 建議先在測試環境中測試流程"]
    ]
    
    # 寫入說明
    for row_idx, instruction in enumerate(instructions, 1):
        if instruction:  # 如果不是空行
            instruction_ws.cell(row=row_idx, column=1, value=instruction[0])
            
            # 設定標題樣式
            cell = instruction_ws.cell(row=row_idx, column=1)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14, color="366092")
            elif instruction[0].startswith("步驟") or instruction[0].endswith("說明：") or instruction[0] == "注意事項：":
                cell.font = Font(bold=True, size=12)
    
    # 設定說明工作表的欄寬
    instruction_ws.column_dimensions['A'].width = 80
    
    # 儲存檔案
    wb.save(output_path)
    
    print(f"✅ Excel 觸發檔案已建立：{output_path}")
    print("\n📋 下一步：")
    print("1. 將此檔案上傳到 OneDrive 或 SharePoint")
    print("2. 在 Power Automate 中更新檔案 ID")
    print("3. 開始使用！")
    
    return output_path


def create_batch_import_template(
    processed_folders: list,
    output_path: str = "batch_permissions_import.xlsx"
):
    """
    建立批次匯入模板，用於大量設定權限
    
    Args:
        processed_folders: 已處理的資料夾清單
        output_path: 輸出檔案路徑
    """
    print(f"\n🚀 建立批次權限匯入檔案...")
    
    # 建立新的工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "批次權限設定"
    
    # 定義欄位標題
    headers = [
        '審查者名稱', '資料夾名稱', '資料夾完整路徑', 'Email', 
        '權限等級', '處理狀態', '處理時間', '處理結果', 
        '站台名稱', '文件庫'
    ]
    
    # 寫入標題
    ws.append(headers)
    
    # 寫入資料
    for folder_info in processed_folders:
        row = [
            folder_info['reviewer'],
            folder_info['folder_name'], 
            folder_info['full_path'],
            f"{folder_info['reviewer'].lower().replace(' ', '.')}@company.com",
            'Contribute',
            '待處理',
            '',
            '',
            'https://company.sharepoint.com/sites/YourSite',
            'Documents'
        ]
        ws.append(row)
    
    # 建立表格
    if ws.max_row > 1:
        tab = Table(displayName="批次權限表", ref=f"A1:J{ws.max_row}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
    
    # 儲存檔案
    wb.save(output_path)
    
    print(f"✅ 批次匯入檔案已建立：{output_path}")
    print(f"📊 包含 {len(processed_folders)} 個資料夾的權限設定")
    
    return output_path


if __name__ == "__main__":
    # 建立觸發檔案
    trigger_file = create_power_automate_trigger_excel()
    
    # 建立批次匯入範例
    sample_folders = [
        {'reviewer': '張三', 'folder_name': '張三', 'full_path': '/Documents/張三'},
        {'reviewer': '李四', 'folder_name': '李四', 'full_path': '/Documents/李四'},
        {'reviewer': '王五', 'folder_name': '王五', 'full_path': '/Documents/王五'}
    ]
    
    batch_file = create_batch_import_template(sample_folders)
    
    print("\n✨ 所有檔案已建立完成！")