#!/usr/bin/env python3
"""
修正版 Excel 分割器 - 使用隱藏列而非刪除列
解決 OneDrive 版本的檔案格式問題
"""

import os
import shutil
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import glob
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import re

def sanitize_folder_name(name: str) -> str:
    """清理資料夾名稱，確保相容性"""
    invalid_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|', '#', '%']
    sanitized = name.strip()
    
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '_')
    
    # 限制長度
    if len(sanitized) > 255:
        sanitized = sanitized[:255].rstrip()
    
    return sanitized

def find_column(worksheet, column_name):
    """在工作表中尋找欄位"""
    for col_idx, cell in enumerate(worksheet[1], start=1):
        if cell.value == column_name:
            return col_idx
    raise ValueError(f"找不到 '{column_name}' 欄位！")

def process_reviewer_excel_hide_rows(file_path, reviewer, column_name, output_folder):
    """
    使用隱藏列方法處理 Excel（保留檔案完整性）
    這是解決檔案格式問題的核心方法
    """
    try:
        # 清理審查者名稱
        reviewer_name = sanitize_folder_name(str(reviewer).strip())
        
        # 建立審查者資料夾
        reviewer_folder = os.path.join(output_folder, reviewer_name)
        os.makedirs(reviewer_folder, exist_ok=True)
        
        # 建立新檔名
        base_name = os.path.basename(file_path)
        name_without_ext = os.path.splitext(base_name)[0]
        ext = os.path.splitext(base_name)[1]
        new_filename = f"{name_without_ext} - {reviewer_name}{ext}"
        dst_path = os.path.join(reviewer_folder, new_filename)
        
        # 先複製整個檔案
        shutil.copy2(file_path, dst_path)
        print(f"  ✓ 已複製檔案: {new_filename}")
        
        # 使用 openpyxl 處理複製的檔案
        wb = load_workbook(dst_path, data_only=False, keep_vba=True, keep_links=True)
        main_ws = wb.active
        
        # 尋找審查者欄位
        col_idx = find_column(main_ws, column_name)
        
        # 隱藏不相關的列（而非刪除）
        rows_to_hide = []
        for row in range(2, main_ws.max_row + 1):
            cell_value = main_ws.cell(row=row, column=col_idx).value
            if str(cell_value) != str(reviewer):
                rows_to_hide.append(row)
        
        print(f"  ✓ 找到 {len(rows_to_hide)} 列需要隱藏")
        
        # 隱藏非相關列
        for row in rows_to_hide:
            main_ws.row_dimensions[row].hidden = True
        
        # 設定自動篩選（可選）
        if main_ws.max_row > 1:
            filter_range = f"A1:{get_column_letter(main_ws.max_column)}{main_ws.max_row}"
            main_ws.auto_filter.ref = filter_range
            
            # 設定篩選條件
            try:
                main_ws.auto_filter.add_filter_column(col_idx - 1, [str(reviewer)])
            except Exception as e:
                print(f"  ⚠️ 無法設定自動篩選: {e}")
        
        # 儲存變更
        wb.save(dst_path)
        wb.close()
        
        print(f"  ✓ 已處理完成，使用隱藏列方法保持檔案完整性")
        
        return True, reviewer_folder, new_filename
        
    except Exception as e:
        print(f"❌ 處理 {reviewer} 的檔案時發生錯誤: {str(e)}")
        return False, None, None

def process_reviewer_excel_minimal_impact(file_path, reviewer, column_name, output_folder):
    """
    最小影響處理方法 - 僅設定篩選，不修改資料結構
    """
    try:
        reviewer_name = sanitize_folder_name(str(reviewer).strip())
        reviewer_folder = os.path.join(output_folder, reviewer_name)
        os.makedirs(reviewer_folder, exist_ok=True)
        
        base_name = os.path.basename(file_path)
        name_without_ext = os.path.splitext(base_name)[0]
        ext = os.path.splitext(base_name)[1]
        new_filename = f"{name_without_ext} - {reviewer_name}{ext}"
        dst_path = os.path.join(reviewer_folder, new_filename)
        
        # 直接複製檔案
        shutil.copy2(file_path, dst_path)
        print(f"  ✓ 已複製檔案: {new_filename}")
        
        # 僅設定篩選，不修改工作表結構
        wb = load_workbook(dst_path, data_only=False, keep_vba=True, keep_links=True)
        main_ws = wb.active
        
        # 尋找審查者欄位
        col_idx = find_column(main_ws, column_name)
        
        # 設定自動篩選範圍
        if main_ws.max_row > 1:
            filter_range = f"A1:{get_column_letter(main_ws.max_column)}{main_ws.max_row}"
            main_ws.auto_filter.ref = filter_range
            
            # 設定篩選條件，只顯示該審查者的資料
            try:
                main_ws.auto_filter.add_filter_column(col_idx - 1, [str(reviewer)])
                print(f"  ✓ 已設定篩選條件顯示 {reviewer} 的資料")
            except Exception as e:
                print(f"  ⚠️ 無法設定篩選條件: {e}")
        
        # 儲存變更
        wb.save(dst_path)
        wb.close()
        
        print(f"  ✓ 已處理完成，保持完整檔案結構")
        
        return True, reviewer_folder, new_filename
        
    except Exception as e:
        print(f"❌ 處理 {reviewer} 的檔案時發生錯誤: {str(e)}")
        return False, None, None

def validate_excel_file(file_path):
    """驗證 Excel 檔案的完整性"""
    try:
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        
        # 基本檢查
        checks = {
            'has_data': ws.max_row > 1,
            'has_columns': ws.max_column > 0,
            'first_row_exists': ws.cell(1, 1).value is not None,
            'no_major_errors': True
        }
        
        wb.close()
        return checks
    except Exception as e:
        print(f"檔案驗證失敗: {e}")
        return {'validation_error': str(e)}

def copy_selected_documents(source_dir, dest_dir, copy_word=True, copy_pdf=True):
    """複製選定的文件類型"""
    copied_files = []
    
    if copy_word:
        word_patterns = [
            os.path.join(source_dir, "*.docx"),
            os.path.join(source_dir, "*.doc")
        ]
        
        for pattern in word_patterns:
            for file in glob.glob(pattern):
                if os.path.isfile(file):
                    dest_path = os.path.join(dest_dir, os.path.basename(file))
                    shutil.copy2(file, dest_path)
                    copied_files.append(os.path.basename(file))
    
    if copy_pdf:
        pdf_pattern = os.path.join(source_dir, "*.pdf")
        for file in glob.glob(pdf_pattern):
            if os.path.isfile(file):
                dest_path = os.path.join(dest_dir, os.path.basename(file))
                shutil.copy2(file, dest_path)
                copied_files.append(os.path.basename(file))
    
    return copied_files

def process_excel_file_safe(file_path, column_name, output_folder, processing_method='hide_rows'):
    """
    安全的 Excel 處理主函數 - 避免檔案格式問題
    
    Args:
        file_path: Excel 檔案路徑
        column_name: 審查者欄位名稱
        output_folder: 輸出資料夾
        processing_method: 處理方法 ('hide_rows', 'filter_only', 'minimal')
    """
    print(f"📁 處理檔案: {os.path.basename(file_path)}")
    print(f"📊 審查者欄位: {column_name}")
    print(f"📂 輸出資料夾: {output_folder}")
    print(f"🔧 處理方法: {processing_method}")
    print("=" * 50)
    
    # 驗證輸入檔案
    if not os.path.exists(file_path):
        print(f"❌ 找不到檔案: {file_path}")
        return False
    
    # 檔案完整性檢查
    validation = validate_excel_file(file_path)
    if 'validation_error' in validation:
        print(f"❌ 檔案驗證失敗: {validation['validation_error']}")
        return False
    
    try:
        # 讀取 Excel 檔案
        df = pd.read_excel(file_path, engine='openpyxl')
        
        if column_name not in df.columns:
            print(f"❌ 找不到欄位 '{column_name}'")
            print(f"可用欄位: {', '.join(df.columns)}")
            return False
        
        # 取得唯一審查者
        reviewers = df[column_name].dropna().unique().tolist()
        print(f"✓ 找到 {len(reviewers)} 位審查者")
        
        # 處理每位審查者
        processed = 0
        failed = 0
        
        for i, reviewer in enumerate(reviewers):
            print(f"\n📝 處理中: {reviewer} ({i+1}/{len(reviewers)})")
            
            # 根據選擇的方法處理
            if processing_method == 'minimal':
                success, folder_path, filename = process_reviewer_excel_minimal_impact(
                    file_path, reviewer, column_name, output_folder
                )
            else:  # 預設使用隱藏列方法
                success, folder_path, filename = process_reviewer_excel_hide_rows(
                    file_path, reviewer, column_name, output_folder
                )
            
            if success:
                # 驗證輸出檔案
                output_file_path = os.path.join(folder_path, filename)
                output_validation = validate_excel_file(output_file_path)
                
                if 'validation_error' in output_validation:
                    print(f"  ⚠️ 輸出檔案驗證失敗: {output_validation['validation_error']}")
                    failed += 1
                else:
                    print(f"  ✓ 輸出檔案驗證通過")
                    processed += 1
            else:
                failed += 1
        
        # 總結
        print("\n" + "=" * 50)
        print(f"✅ 處理完成！")
        print(f"📊 成功處理: {processed}/{len(reviewers)} 位審查者")
        if failed > 0:
            print(f"❌ 處理失敗: {failed} 位")
        print(f"📁 輸出位置: {output_folder}")
        
        return processed > 0
        
    except Exception as e:
        print(f"\n❌ 發生錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

# 測試函數
def test_processing_methods():
    """測試不同的處理方法"""
    print("測試不同處理方法的相容性...")
    
    methods = [
        ('hide_rows', '隱藏列方法（推薦）'),
        ('minimal', '最小影響方法（最安全）')
    ]
    
    for method, description in methods:
        print(f"\n{method}: {description}")
        print("  優點: 保持檔案完整性，避免格式問題")
        print("  缺點: 檔案大小不會減少")

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 3:
        print("使用方式: python excel_splitter_fixed.py <Excel檔案> <審查者欄位> [輸出資料夾] [處理方法]")
        print("範例: python excel_splitter_fixed.py data.xlsx Reviewer ./output hide_rows")
        print("\n處理方法:")
        test_processing_methods()
        sys.exit(1)
    
    file_path = sys.argv[1]
    column_name = sys.argv[2]
    output_folder = sys.argv[3] if len(sys.argv) > 3 else os.path.dirname(file_path)
    method = sys.argv[4] if len(sys.argv) > 4 else 'hide_rows'
    
    success = process_excel_file_safe(file_path, column_name, output_folder, method)
    sys.exit(0 if success else 1)